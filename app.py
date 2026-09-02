"""
Transformador de Fichas de Clientes
====================================
Aplicacao web local para transformar ficheiros Excel de clientes
no formato estruturado 'clientes_final'.

Abrir com duplo clique no ficheiro 'Abrir Transformador.bat'
ou executar: python app.py
"""

import os
import sys
import re
import io
import webbrowser
import threading
from datetime import datetime

import pandas as pd
from flask import Flask, request, send_file, render_template_string, jsonify
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ─── HTML Template ────────────────────────────────────────────────────────────

HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="pt">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Transformador de Fichas de Clientes</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }

        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            color: #fff;
        }

        .container {
            max-width: 700px;
            width: 90%;
            text-align: center;
        }

        .logo {
            font-size: 3rem;
            margin-bottom: 0.5rem;
            animation: pulse 2s ease-in-out infinite;
        }

        @keyframes pulse {
            0%, 100% { transform: scale(1); }
            50% { transform: scale(1.05); }
        }

        h1 {
            font-size: 1.8rem;
            font-weight: 300;
            margin-bottom: 0.3rem;
            letter-spacing: 1px;
        }

        .subtitle {
            color: #a0a0c0;
            font-size: 0.95rem;
            margin-bottom: 2rem;
        }

        /* Drop Zone */
        .drop-zone {
            border: 3px dashed rgba(255, 255, 255, 0.25);
            border-radius: 20px;
            padding: 60px 40px;
            transition: all 0.3s ease;
            cursor: pointer;
            background: rgba(255, 255, 255, 0.03);
            position: relative;
        }

        .drop-zone:hover,
        .drop-zone.dragover {
            border-color: #7c5cfc;
            background: rgba(124, 92, 252, 0.08);
            transform: scale(1.02);
        }

        .drop-zone-icon {
            font-size: 4rem;
            margin-bottom: 1rem;
            opacity: 0.7;
        }

        .drop-zone-text {
            font-size: 1.2rem;
            color: #c0c0e0;
            margin-bottom: 0.5rem;
        }

        .drop-zone-hint {
            font-size: 0.85rem;
            color: #808090;
        }

        .drop-zone input[type="file"] {
            display: none;
        }

        /* Processing State */
        .processing {
            display: none;
            padding: 60px 40px;
        }

        .spinner {
            width: 60px;
            height: 60px;
            border: 4px solid rgba(255, 255, 255, 0.1);
            border-left-color: #7c5cfc;
            border-radius: 50%;
            animation: spin 0.8s linear infinite;
            margin: 0 auto 1.5rem;
        }

        @keyframes spin {
            to { transform: rotate(360deg); }
        }

        .processing-text {
            font-size: 1.1rem;
            color: #c0c0e0;
        }

        /* Result State */
        .result {
            display: none;
            padding: 40px;
        }

        .result-success {
            background: rgba(76, 175, 80, 0.1);
            border: 2px solid rgba(76, 175, 80, 0.3);
            border-radius: 20px;
            padding: 40px;
        }

        .result-icon {
            font-size: 4rem;
            margin-bottom: 1rem;
        }

        .result-title {
            font-size: 1.4rem;
            font-weight: 600;
            margin-bottom: 0.5rem;
            color: #4caf50;
        }

        .result-stats {
            display: flex;
            justify-content: center;
            gap: 2rem;
            margin: 1.5rem 0;
            flex-wrap: wrap;
        }

        .stat {
            text-align: center;
        }

        .stat-value {
            font-size: 2rem;
            font-weight: 700;
            color: #7c5cfc;
        }

        .stat-label {
            font-size: 0.8rem;
            color: #808090;
            text-transform: uppercase;
            letter-spacing: 1px;
        }

        .btn {
            display: inline-block;
            padding: 14px 40px;
            border-radius: 50px;
            font-size: 1rem;
            font-weight: 600;
            text-decoration: none;
            transition: all 0.3s ease;
            cursor: pointer;
            border: none;
            margin: 0.5rem;
        }

        .btn-download {
            background: linear-gradient(135deg, #7c5cfc, #5a3fd4);
            color: #fff;
            box-shadow: 0 4px 20px rgba(124, 92, 252, 0.4);
        }

        .btn-download:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 30px rgba(124, 92, 252, 0.6);
        }

        .btn-reset {
            background: rgba(255, 255, 255, 0.1);
            color: #c0c0e0;
        }

        .btn-reset:hover {
            background: rgba(255, 255, 255, 0.15);
        }

        /* Error State */
        .result-error {
            background: rgba(244, 67, 54, 0.1);
            border: 2px solid rgba(244, 67, 54, 0.3);
            border-radius: 20px;
            padding: 40px;
        }

        .result-error .result-title {
            color: #f44336;
        }

        .error-message {
            color: #e0a0a0;
            font-size: 0.9rem;
            margin-top: 1rem;
            word-break: break-word;
        }

        /* Footer */
        .footer {
            margin-top: 2rem;
            color: #505060;
            font-size: 0.75rem;
        }

        /* Table Preview */
        .preview-table {
            margin: 1.5rem auto;
            max-width: 100%;
            overflow-x: auto;
        }

        .preview-table table {
            border-collapse: collapse;
            font-size: 0.75rem;
            width: 100%;
        }

        .preview-table th {
            background: rgba(124, 92, 252, 0.2);
            color: #c0c0e0;
            padding: 8px 12px;
            text-align: left;
            white-space: nowrap;
        }

        .preview-table td {
            padding: 6px 12px;
            border-bottom: 1px solid rgba(255,255,255,0.05);
            color: #a0a0b0;
            white-space: nowrap;
        }

        .preview-table tr:hover td {
            background: rgba(255,255,255,0.03);
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="logo">&#128203;</div>
        <h1>Transformador de Fichas</h1>
        <p class="subtitle">Arrasta o ficheiro Excel e ele fica estruturado em segundos</p>

        <!-- DROP ZONE -->
        <div class="drop-zone" id="dropZone">
            <div class="drop-zone-icon">&#128194;</div>
            <div class="drop-zone-text">Arrasta o ficheiro Excel para aqui</div>
            <div class="drop-zone-hint">ou clica para selecionar o ficheiro (.xlsx)</div>
            <input type="file" id="fileInput" accept=".xlsx,.xls" multiple>
        </div>
        
        <div style="margin-top: 15px; margin-bottom: 5px;" id="mergeOption">
            <label style="cursor: pointer; font-size: 0.95rem; color: #a0a0c0; display: flex; align-items: center; justify-content: center; gap: 8px;">
                <input type="checkbox" id="mergeCheckbox" style="width: 16px; height: 16px; cursor: pointer;">
                Juntar todos os ficheiros num único Excel
            </label>
        </div>

        <!-- PROCESSING -->
        <div class="processing" id="processing">
            <div class="spinner"></div>
            <div class="processing-text">A transformar ficheiro(s)...</div>
        </div>

        <!-- RESULT -->
        <div class="result" id="result"></div>

        <div class="footer">
            Equipa Red Team Alcino Fontes &middot; Transformador de Fichas v2.0
        </div>
    </div>

    <script>
        const dropZone = document.getElementById('dropZone');
        const fileInput = document.getElementById('fileInput');
        const processing = document.getElementById('processing');
        const result = document.getElementById('result');

        // Click to select
        dropZone.addEventListener('click', () => fileInput.click());

        // Drag events
        dropZone.addEventListener('dragover', (e) => {
            e.preventDefault();
            dropZone.classList.add('dragover');
        });

        dropZone.addEventListener('dragleave', () => {
            dropZone.classList.remove('dragover');
        });

        dropZone.addEventListener('drop', (e) => {
            e.preventDefault();
            dropZone.classList.remove('dragover');
            const files = e.dataTransfer.files;
            if (files.length > 0) processFiles(files);
        });

        // File input change
        fileInput.addEventListener('change', () => {
            if (fileInput.files.length > 0) processFiles(fileInput.files);
        });

        function processFiles(files) {
            // Validate
            let validFiles = [];
            for (let i = 0; i < files.length; i++) {
                if (files[i].name.match(/\.xlsx?$/i)) {
                    validFiles.push(files[i]);
                }
            }

            if (validFiles.length === 0) {
                showError('Por favor seleciona ficheiro(s) Excel (.xlsx)');
                return;
            }

            // Show processing
            dropZone.style.display = 'none';
            document.getElementById('mergeOption').style.display = 'none';
            processing.style.display = 'block';
            result.style.display = 'none';

            // Upload
            const formData = new FormData();
            for (let i = 0; i < validFiles.length; i++) {
                formData.append('file', validFiles[i]);
            }
            
            const mergeCheckbox = document.getElementById('mergeCheckbox');
            if (mergeCheckbox && mergeCheckbox.checked) {
                formData.append('merge', 'true');
            }

            fetch('/transformar', {
                method: 'POST',
                body: formData
            })
            .then(response => {
                if (!response.ok) {
                    return response.json().then(data => { throw new Error(data.error); });
                }
                return response.json();
            })
            .then(data => {
                showSuccess(data, validFiles.length);
            })
            .catch(error => {
                showError(error.message || 'Erro ao processar ficheiros');
            });
        }

        function showSuccess(data, numFiles) {
            processing.style.display = 'none';
            result.style.display = 'block';

            let previewHtml = '';
            if (data.preview && data.preview.length > 0) {
                const cols = Object.keys(data.preview[0]);
                previewHtml = `
                    <div class="preview-table">
                        <table>
                            <thead><tr>${cols.map(c => `<th>${c}</th>`).join('')}</tr></thead>
                            <tbody>${data.preview.map(row =>
                                `<tr>${cols.map(c => `<td>${row[c] !== null && row[c] !== undefined ? row[c] : ''}</td>`).join('')}</tr>`
                            ).join('')}</tbody>
                        </table>
                    </div>
                `;
            }

            const title = numFiles > 1 ? `${numFiles} Ficheiros Transformados!` : 'Ficheiro Transformado!';
            const btnText = data.is_zip ? '&#11015; Descarregar ZIP com Todos' : '&#11015; Descarregar Excel';

            result.innerHTML = `
                <div class="result-success">
                    <div class="result-icon">&#9989;</div>
                    <div class="result-title">${title}</div>
                    <div class="result-stats">
                        <div class="stat">
                            <div class="stat-value">${data.total_linhas}</div>
                            <div class="stat-label">Linhas Origem</div>
                        </div>
                        <div class="stat">
                            <div class="stat-value">${data.processados}</div>
                            <div class="stat-label">Clientes</div>
                        </div>
                        <div class="stat">
                            <div class="stat-value">${data.ignorados}</div>
                            <div class="stat-label">Sem Telefone</div>
                        </div>
                    </div>
                    ${previewHtml}
                    <a href="/descarregar/${data.ficheiro}" class="btn btn-download">${btnText}</a>
                    <button class="btn btn-reset" onclick="resetForm()">&#128260; Novo(s) Ficheiro(s)</button>
                </div>
            `;
        }

        function showError(message) {
            processing.style.display = 'none';
            result.style.display = 'block';
            result.innerHTML = `
                <div class="result-error">
                    <div class="result-icon">&#10060;</div>
                    <div class="result-title">Erro ao Processar</div>
                    <div class="error-message">${message}</div>
                    <br>
                    <button class="btn btn-reset" onclick="resetForm()">&#128260; Tentar Novamente</button>
                </div>
            `;
        }

        function resetForm() {
            dropZone.style.display = 'block';
            document.getElementById('mergeOption').style.display = 'block';
            processing.style.display = 'none';
            result.style.display = 'none';
            fileInput.value = '';
        }
    </script>
</body>
</html>
"""


# ─── Logica de Transformacao ──────────────────────────────────────────────────

def extrair_nome(texto):
    """Extrai o nome (primeira parte antes do primeiro ;)."""
    parts = texto.split(';')
    nome = parts[0].strip()
    return nome if nome else None


def extrair_telefones(texto):
    """Extrai telefones validos do final do texto."""
    parts = texto.split(';')
    phone_part = parts[-1].strip() if parts else ''
    phone_part = phone_part.replace('\n', '').strip()

    phones = [p.strip() for p in phone_part.split(',')]
    valid_phones = []
    for p in phones:
        digits = re.sub(r'\D', '', p)
        if len(digits) >= 9:
            valid_phones.append(int(digits))
    return valid_phones


def transformar_dataframe(df_input):
    """Transforma o DataFrame de entrada no formato clientes_final, sem apagar linhas."""
    registos = []
    ignorados = 0

    for i in range(len(df_input)):
        celula_0 = df_input.iloc[i, 0]
        texto = str(celula_0) if pd.notna(celula_0) else ""
        data_fid = df_input.iloc[i, 1] if df_input.shape[1] > 1 and pd.notna(df_input.iloc[i, 1]) else None

        nome = extrair_nome(texto) if texto else ""
        telefones = extrair_telefones(texto) if texto else []

        if not telefones:
            ignorados += 1

        registos.append({
            'Nome': nome,
            'Telefone': telefones[0] if len(telefones) > 0 else None,
            'CP7': None,
            'Localidade': None,
            'Morada': None,
            'Operador Atual': None,
            u'Data Fideliza\u00e7\u00e3o': data_fid,
            'Telefone 2': telefones[1] if len(telefones) > 1 else None,
            'Telefone 3': telefones[2] if len(telefones) > 2 else None,
            'Telefone 4': telefones[3] if len(telefones) > 3 else None,
        })

    colunas = ['Nome', 'Telefone', 'CP7', 'Localidade', 'Morada',
               'Operador Atual', u'Data Fideliza\u00e7\u00e3o',
               'Telefone 2', 'Telefone 3', 'Telefone 4']

    df_final = pd.DataFrame(registos, columns=colunas)

    # Tipos corretos
    df_final['Telefone'] = df_final['Telefone'].astype('Int64')
    for col in ['Telefone 2', 'Telefone 3', 'Telefone 4']:
        df_final[col] = df_final[col].astype('Int64')

    return df_final, ignorados


def formatar_excel(caminho_ficheiro, df):
    """Aplica formatacao profissional ao ficheiro Excel gerado."""
    from openpyxl import load_workbook

    wb = load_workbook(caminho_ficheiro)
    ws = wb.active
    ws.title = "Clientes"

    # ─── Estilos ──────────────────────────────────────────────────────────
    # Cor do cabecalho
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Linhas alternadas
    row_fill_even = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Fontes de dados
    data_font = Font(name="Calibri", size=10, color="1A1A2E")
    data_font_phone = Font(name="Calibri", size=10, color="1A1A2E", bold=False)
    data_font_name = Font(name="Calibri", size=10, color="1A1A2E", bold=True)

    # Alinhamentos
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    # Bordas
    thin_border = Border(
        left=Side(style="thin", color="B0C4DE"),
        right=Side(style="thin", color="B0C4DE"),
        top=Side(style="thin", color="B0C4DE"),
        bottom=Side(style="thin", color="B0C4DE"),
    )

    header_border = Border(
        left=Side(style="thin", color="0D3B66"),
        right=Side(style="thin", color="0D3B66"),
        top=Side(style="medium", color="0D3B66"),
        bottom=Side(style="medium", color="0D3B66"),
    )

    # ─── Cabecalho (linha 1) ──────────────────────────────────────────────
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    # ─── Altura do cabecalho ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 30

    # ─── Colunas de telefone e data (indices 1-based) ─────────────────────
    col_names = [cell.value for cell in ws[1]]
    phone_cols = set()
    date_cols = set()
    name_col = None
    for idx, name in enumerate(col_names, 1):
        if name and 'Telefone' in str(name):
            phone_cols.add(idx)
        if name and 'Data' in str(name):
            date_cols.add(idx)
        if name and name == 'Nome':
            name_col = idx

    # ─── Formatar dados (linhas 2+) ──────────────────────────────────────
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), 2):
        # Cor alternada
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        is_even = row_idx % 2 == 0

        # Altura da linha
        ws.row_dimensions[row_idx].height = 22

        for col_idx, cell in enumerate(row, 1):
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = align_left

            if col_idx == name_col:
                # Nome: bold, alinhado a esquerda
                cell.font = data_font_name
                cell.alignment = align_left
            elif col_idx in phone_cols:
                # Telefones: formato texto para nao perder zeros
                cell.font = data_font_phone
                cell.alignment = align_center
                cell.number_format = '0'
            elif col_idx in date_cols:
                # Datas: formato DD/MM/AAAA
                cell.font = data_font
                cell.alignment = align_center
                cell.number_format = 'DD/MM/YYYY'
            else:
                cell.font = data_font
                cell.alignment = align_center

    # ─── Auto-ajustar largura das colunas ─────────────────────────────────
    min_widths = {
        'Nome': 38,
        'Telefone': 14,
        'CP7': 10,
        'Localidade': 18,
        'Morada': 30,
        'Operador Atual': 16,
        'Telefone 2': 14,
        'Telefone 3': 14,
        'Telefone 4': 14,
    }
    # Any col with "Data" in name
    for name in col_names:
        if name and 'Data' in str(name):
            min_widths[name] = 18

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        col_name = col_names[col_idx - 1] if col_idx - 1 < len(col_names) else ''

        # Calcular largura maxima do conteudo
        max_len = len(str(col_name or '')) + 2
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 200), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    cell_len = len(str(cell.value)) + 2
                    max_len = max(max_len, cell_len)

        # Aplicar largura minima definida ou calculada
        min_w = min_widths.get(str(col_name), 12)
        final_width = max(min_w, min(max_len, 50))
        ws.column_dimensions[col_letter].width = final_width

    # ─── Congelar painel (fixar cabecalho) ────────────────────────────────
    ws.freeze_panes = 'A2'

    # ─── Filtro automatico ────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ─── Guardar ──────────────────────────────────────────────────────────
    wb.save(caminho_ficheiro)
    wb.close()


# ─── Pasta temporaria para ficheiros gerados ──────────────────────────────────

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_ficheiros_gerados')
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ─── Rotas Flask ──────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/transformar', methods=['POST'])
def transformar():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum ficheiro enviado'}), 400

    files = request.files.getlist('file')
    if not files or files[0].filename == '':
        return jsonify({'error': 'Nenhum ficheiro selecionado'}), 400

    merge = request.form.get('merge') == 'true'

    try:
        import zipfile
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        total_linhas_all = 0
        processados_all = 0
        ignorados_all = 0
        previews = []

        if merge:
            # Junta todos os ficheiros num unico DataFrame
            all_df_finals = []
            for file in files:
                df_input = pd.read_excel(file, header=None)
                total_linhas_all += len(df_input)
                
                df_final, ignorados = transformar_dataframe(df_input)
                processados_all += len(df_final)
                ignorados_all += ignorados
                all_df_finals.append(df_final)

            df_final_merged = pd.concat(all_df_finals, ignore_index=True) if all_df_finals else pd.DataFrame()
            
            nome_saida = f'lote_clientes_juntos_{timestamp}.xlsx'
            caminho_saida = os.path.join(OUTPUT_DIR, nome_saida)
            
            df_final_merged.to_excel(caminho_saida, index=False, engine='openpyxl')
            formatar_excel(caminho_saida, df_final_merged)

            # Preview das primeiras linhas
            preview_df = df_final_merged.head(5).copy()
            date_col = u'Data Fideliza\u00e7\u00e3o'
            if date_col in preview_df.columns:
                def format_date(x):
                    if pd.isna(x): return ''
                    if hasattr(x, 'strftime'): return x.strftime('%Y-%m-%d')
                    return str(x)
                preview_df[date_col] = preview_df[date_col].apply(format_date)
            previews = preview_df.where(pd.notnull(preview_df), None).to_dict('records')

            ficheiro_final = nome_saida
            is_zip = False

        else:
            # Comportamento normal: ficheiros separados num ZIP
            processed_files = []
            for file in files:
                # Ler o ficheiro Excel
                df_input = pd.read_excel(file, header=None)
                total_linhas = len(df_input)
                total_linhas_all += total_linhas

                # Transformar
                df_final, ignorados = transformar_dataframe(df_input)
                processados = len(df_final)
                
                processados_all += processados
                ignorados_all += ignorados

                # Gerar nome do ficheiro de saida
                safe_name = "".join(c for c in file.filename if c.isalnum() or c in (' ', '.', '_')).rstrip()
                nome_saida = f'formatado_{safe_name}'
                caminho_saida = os.path.join(OUTPUT_DIR, nome_saida)

                # Exportar e formatar
                df_final.to_excel(caminho_saida, index=False, engine='openpyxl')
                formatar_excel(caminho_saida, df_final)
                
                processed_files.append((nome_saida, caminho_saida))

                # Preview das primeiras linhas
                if len(previews) < 5:
                    preview_df = df_final.head(2).copy()
                    date_col = u'Data Fideliza\u00e7\u00e3o'
                    if date_col in preview_df.columns:
                        def format_date(x):
                            if pd.isna(x): return ''
                            if hasattr(x, 'strftime'): return x.strftime('%Y-%m-%d')
                            return str(x)
                        preview_df[date_col] = preview_df[date_col].apply(format_date)
                    previews.extend(preview_df.where(pd.notnull(preview_df), None).to_dict('records'))

            # Se houver mais de um ficheiro, criar ZIP
            if len(processed_files) > 1:
                zip_filename = f'lote_clientes_{timestamp}.zip'
                zip_path = os.path.join(OUTPUT_DIR, zip_filename)
                with zipfile.ZipFile(zip_path, 'w') as zipf:
                    for nome_saida, caminho_saida in processed_files:
                        zipf.write(caminho_saida, arcname=nome_saida)
                ficheiro_final = zip_filename
                is_zip = True
            else:
                ficheiro_final = processed_files[0][0]
                is_zip = False

        return jsonify({
            'total_linhas': total_linhas_all,
            'processados': processados_all,
            'ignorados': ignorados_all,
            'ficheiro': ficheiro_final,
            'is_zip': is_zip,
            'preview': previews[:5]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/descarregar/<nome_ficheiro>')
def descarregar(nome_ficheiro):
    caminho = os.path.join(OUTPUT_DIR, nome_ficheiro)
    if not os.path.exists(caminho):
        return 'Ficheiro nao encontrado', 404
    return send_file(
        caminho,
        as_attachment=True,
        download_name=nome_ficheiro,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ─── Main ─────────────────────────────────────────────────────────────────────

def open_browser():
    """Abre o browser automaticamente apos o servidor iniciar."""
    webbrowser.open('http://127.0.0.1:5000')


if __name__ == '__main__':
    import socket

    # Descobrir o IP local da maquina
    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except Exception:
            return "127.0.0.1"

    local_ip = get_local_ip()
    port = 5000

    print('=' * 55)
    print('  Transformador de Fichas de Clientes')
    print('=' * 55)
    print(f'  Este computador:  http://127.0.0.1:{port}')
    print(f'  Rede local:       http://{local_ip}:{port}')
    print('=' * 55)
    print('  NAO feche esta janela enquanto usar!')
    print('=' * 55)

    # Abrir browser automaticamente
    threading.Timer(1.5, open_browser).start()

    app.run(host='0.0.0.0', port=port, debug=False)
